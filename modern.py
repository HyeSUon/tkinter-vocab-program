import customtkinter
import os
import csv
import random
import tkinter as tk
from tkinter import ttk
from PIL import Image
from tkinter import filedialog
from tkinter import messagebox
from customScrollableFrame import *
from openpyxl.styles import Font
from openpyxl import Workbook

class App(customtkinter.CTk):
    def __init__(self):
        super().__init__()


        self.app_data = {"dict": {},
                         "double_dict": {},
                         "question_count": 0,
                         "word_count": 0,
                         "var_vocab": tk.IntVar(),
                         "word_shuffle": tk.IntVar(),
                         "question_list": [],
                         "name_list": [],
                         "entries": [],
                         "result_list": [],
                         'answer_list': [],
        }
        self.title("영단어 Test")
        self.geometry("800x450")

        # set grid layout 1x2
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)

        # load images with light and dark mode image
        image_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), "images")
        self.logo_image = customtkinter.CTkImage(Image.open(os.path.join(image_path, "CustomTkinter_logo_single.png")), size=(26, 26))
        self.home_image = customtkinter.CTkImage(light_image=Image.open(os.path.join(image_path, "add_folder_dark.png")),
                                                 dark_image=Image.open(os.path.join(image_path, "add_folder_light.png")), size=(20, 20))
        self.chat_image = customtkinter.CTkImage(light_image=Image.open(os.path.join(image_path, "quiz_light.png")),
                                                 dark_image=Image.open(os.path.join(image_path, "quiz_dark.png")), size=(20, 20))
        self.add_user_image = customtkinter.CTkImage(light_image=Image.open(os.path.join(image_path, "result_dark.png")),
                                                     dark_image=Image.open(os.path.join(image_path, "result_light.png")), size=(20, 20))



        # create navigation frame
        self.navigation_frame = customtkinter.CTkFrame(self, corner_radius=0)
        self.navigation_frame.grid(row=0, column=0, sticky="nsew")
        self.navigation_frame.grid_rowconfigure(4, weight=1)

        self.navigation_frame_label = customtkinter.CTkLabel(self.navigation_frame, text="  영단어 Test", image=self.logo_image,
                                                             compound="left", font=customtkinter.CTkFont(size=15, weight="bold"))
        self.navigation_frame_label.grid(row=0, column=0, padx=20, pady=20)

        self.home_button = customtkinter.CTkButton(self.navigation_frame, corner_radius=0, height=40, border_spacing=10, text="Setting",
                                                   fg_color="transparent", text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"),
                                                   image=self.home_image, anchor="w", command=self.home_button_event)
        self.home_button.grid(row=1, column=0, sticky="ew")

        self.frame_2_button = customtkinter.CTkButton(self.navigation_frame, corner_radius=0, height=40, border_spacing=10, text="Test",
                                                      fg_color="transparent", text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"),
                                                      image=self.chat_image, anchor="w", command=self.frame_2_button_event)
        self.frame_2_button.grid(row=2, column=0, sticky="ew")

        self.frame_3_button = customtkinter.CTkButton(self.navigation_frame, corner_radius=0, height=40, border_spacing=10, text="Result",
                                                      fg_color="transparent", text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"),
                                                      image=self.add_user_image, anchor="w", command=self.frame_3_button_event)
        self.frame_3_button.grid(row=3, column=0, sticky="ew")

        self.appearance_mode_menu = customtkinter.CTkOptionMenu(self.navigation_frame, values=["Light", "Dark", "System"],
                                                                command=self.change_appearance_mode_event)
        self.appearance_mode_menu.grid(row=6, column=0, padx=20, pady=20, sticky="s")

        # create home frame
        self.home_frame = customtkinter.CTkFrame(self, corner_radius=0, fg_color="transparent")

        # configure grid layout 3x3
        self.home_frame.grid_columnconfigure(0, weight=2)
        self.home_frame.grid_columnconfigure((1,2), weight=0)
        self.home_frame.grid_rowconfigure((0, 1, 2), weight=1)

        self.frm_textbox = customtkinter.CTkFrame(self.home_frame)
        self.frm_textbox.grid(row=0, column=0, rowspan=2, sticky="nsew")
        self.frm_textbox.grid_columnconfigure((0,1), weight=1)
        self.frm_textbox.grid_rowconfigure((0,1), weight=1)
        self.textbox = customtkinter.CTkTextbox(self.frm_textbox, width=400, height=300)
        self.textbox.grid(row=0, column=0, padx=(10, 10), pady=(10, 10), sticky="nsew")
        self.btn_openfile = customtkinter.CTkButton(self.frm_textbox, text="Open File", command=self.open_file)
        self.btn_openfile.grid(row=1, column=0, padx=(30,30), pady=(30,30), sticky="n")

        self.frm_buttons = customtkinter.CTkFrame(self.home_frame)
        self.frm_buttons.grid(row=0, column=1, sticky= 'nsew', padx=(20,20),pady=(20,20))
        self.lbl_buttons = customtkinter.CTkLabel(self.frm_buttons, text='문제 유형 선택', width=100)
        self.lbl_buttons.grid(row=0, column=0, padx=(20,20), pady=(10,0))
        

        self.radio_button_1 = customtkinter.CTkRadioButton(self.frm_buttons, text="뜻 빈칸", variable=self.app_data["var_vocab"], value=0)
        self.radio_button_1.grid(row=1, column=0, padx=20, pady=10)
        self.radio_button_2 = customtkinter.CTkRadioButton(self.frm_buttons, text="영어 빈칸", variable=self.app_data["var_vocab"], value=1)
        self.radio_button_2.grid(row=2, column=0, padx=20, pady=10)
        self.radio_button_3 = customtkinter.CTkRadioButton(self.frm_buttons, text="랜덤", variable=self.app_data["var_vocab"], value=2)
        self.radio_button_3.grid(row=3, column=0, padx=20, pady=10)

        self.frm_options = customtkinter.CTkFrame(self.home_frame)
        self.frm_options.grid(row=1, column=1, sticky='nsew', padx=(20,20), pady=(20, 20))
        self.lbl_options = customtkinter.CTkLabel(self.frm_options, text='옵션')
        self.lbl_options.grid(row=0, column=0, pady=10)
        self.checkbox = customtkinter.CTkCheckBox(self.frm_options, text="단어 섞기", variable=self.app_data["word_shuffle"])
        self.checkbox.grid(row=1, column=0, padx=10, pady=10, sticky='s')
        self.btn_option = customtkinter.CTkButton(self.frm_options, text='문제 수 설정', command=self.open_input_dialog_event)
        self.btn_option.grid(row=2, column=0, padx=10, pady=10)

        # create second frame
        self.second_frame = customtkinter.CTkFrame(self, corner_radius=0, fg_color="transparent")

        # 스크롤바
        self.frm_scroll = ScrollableFrame(self.second_frame)
        self.frm_scroll.pack(expand=True, fill='y', pady=10)

        #제출하기 버튼
        self.btn_enter = ctk.CTkButton(self.second_frame, text="Submit", width=10, command=self.scoring)
        self.btn_enter.pack()


        # create third frame
        self.third_frame = customtkinter.CTkFrame(self, corner_radius=0, fg_color="transparent")

        # configure grid layout 2x1
        self.frm_result = customtkinter.CTkFrame(self.third_frame)
        self.frm_result.pack(fill ='both', pady=30)
        self.btn_saveas = customtkinter.CTkButton(self.frm_result, text="Save as", command=self.save)
        self.btn_saveas.pack(side='right', anchor='s', pady=30, padx=10)
        self.resultbox = customtkinter.CTkTextbox(self.frm_result, width=450, height=400)
        self.resultbox.pack(side='right',pady=10)

        # select default frame
        self.select_frame_by_name("home")



    def scoring(self):
        result = messagebox.askquestion("Submit", "정말로 제출하시겠습니까?")
        answer = self.app_data["double_dict"]
        result_list = self.app_data["result_list"]
        question_list = self.app_data["question_list"]
        answer_list = self.app_data['answer_list']
        result_list.clear()
        if result == 'yes':
            if self.app_data['word_count'] == 0 or self.app_data['question_count'] == 0:
                messagebox.showerror("경고", "문제 수가 충분하지 않습니다.")
                return
            answer_list.clear()
            entries = self.app_data["entries"]
            for idx,entry in enumerate(entries):
                if idx >= self.app_data["question_count"]:
                    break
                answer_list.append(entry.get())
                try:
                    if answer[question_list[idx]] == entry.get():
                        result_list.append("O")
                    else:
                        result_list.append("X")
                except:
                    result_list.append("X")
            self.select_frame_by_name('frame_3')
        else:
            return
        self.resultUpdate()

    def nextWidget(self, event):
        event.widget.tk_focusNext().focus()
        return("break")
        
    def select_frame_by_name(self, name):
        # set button color for selected button
        self.home_button.configure(fg_color=("gray75", "gray25") if name == "home" else "transparent")
        self.frame_2_button.configure(fg_color=("gray75", "gray25") if name == "frame_2" else "transparent")
        self.frame_3_button.configure(fg_color=("gray75", "gray25") if name == "frame_3" else "transparent")

        # show selected frame
        if name == "home":
            self.home_frame.grid(row=0, column=1, sticky="nsew")
        else:
            self.home_frame.grid_forget()
        if name == "frame_2":
            self.second_frame.grid(row=0, column=1, sticky="nsew")
            self.frm2_update()
        else:
            self.second_frame.grid_forget()
        if name == "frame_3":
            self.third_frame.grid(row=0, column=1, sticky="nsew")
        else:
            self.third_frame.grid_forget()

    def save(self):
        result_list = self.app_data['result_list']

        if len(result_list) == 0:
            messagebox.showerror("경고", "문제를 제출하고 눌러주세요.")
            return

        wb = Workbook()
        sheet = wb.active
        question_list = self.app_data["question_list"]
        result_list = self.app_data['result_list']
        answer_list = self.app_data['answer_list']
        double_dict = self.app_data['double_dict']

        sheet.cell(row=1, column=1).value ='문제'
        sheet.cell(row=1, column=3).value = '입력'
        sheet.cell(row=1, column=5).value = '정답'

        for i in range(self.app_data['question_count']):
            sheet.cell(row = i+2, column= 1).value= str(i+1) + ". " + question_list[i]
            if result_list[i] == 'O':
                sheet.cell(row=i+2, column=3).font = Font(color='008000')
                sheet.cell(row= i+2, column=3).value = answer_list[i]
            else:
                sheet.cell(row=i+2, column=3).font = Font(color='FF0000', strike=True, bold=True)
                if answer_list[i] == '':
                    sheet.cell(row= i+2, column=3).value = "------"
                else:
                    sheet.cell(row= i+2, column=3).value = answer_list[i]
                sheet.cell(row=i+2, column=5).font = Font(color='FF0000', bold=True)
                sheet.cell(row=i+2, column=5).value = double_dict[question_list[i]]
        filename = filedialog.asksaveasfilename(initialdir='/', title="Select folder",
                                                initialfile='result',
                                                defaultextension='.csv',
                                                filetypes=(("CSV files", "*.csv"),
                                                ("XLSX files", "*.xlsx")))
        if filename is None or filename == '':
            return
        try:
            wb.save(filename)
        except:
            messagebox.showerror("에러", "파일을 저장할 수 없습니다 \n(저장하려는 파일이 사용중인지 확인하세요.)")
            return
        messagebox.showinfo("정보", "저장 되었습니다.")


    def resultUpdate(self):

        question_list = self.app_data["question_list"] 
        result_list = self.app_data["result_list"]
        double_dict = self.app_data["double_dict"]
        entries = self.app_data['entries']

        self.resultbox.delete(0.0, tk.END)
        self.resultbox.insert(tk.END, '---------------------------------------------------------------------------------------------------------------------------------------\n')
        self.resultbox.insert(tk.END, "총 문제 수: " + str(self.app_data['question_count']) + '\n')
        self.resultbox.insert(tk.END, "맞힌 수: ")
        count = 0
        for i in range(self.app_data['question_count']):
            if result_list[i] == 'O':
                count += 1
        self.resultbox.insert(tk.END, str(count) + '\n')
        self.resultbox.insert(tk.END, "틀린 수: ")
        self.resultbox.insert(tk.END, str(self.app_data["question_count"] - count) + '\n')
        self.resultbox.insert(tk.END, '---------------------------------------------------------------------------------------------------------------------------------------\n')
        self.resultbox.insert(tk.END, '오답노트를 보려면 Save as... 버튼을 클릭하세요.')

    def open_input_dialog_event(self):
        self.dialog = customtkinter.CTkInputDialog(text="출제 수(max:"+ str(self.app_data["word_count"])+"): ", title="옵션", insert_text=str(self.app_data['question_count']))
        try:
            num = int(self.dialog.get_input())
            if num < 0 or num > self.app_data['word_count']:
                return
            self.app_data['question_count'] = num
        except:
            return

    def home_button_event(self):
        self.select_frame_by_name("home")

    def frame_2_button_event(self):
        self.select_frame_by_name("frame_2")

    def frame_3_button_event(self):
        self.select_frame_by_name("frame_3")

    def change_appearance_mode_event(self, new_appearance_mode):
        customtkinter.set_appearance_mode(new_appearance_mode)
        if new_appearance_mode == 'Light':
            self.frm_scroll.canvas['bg'] = '#dbdbdb'
        else:
            self.frm_scroll.canvas['bg'] = '#2b2b2b'
    
    def frm2_update(self):
        question_list = self.app_data["question_list"]
        question_list.clear()
        var_vocab = self.app_data["var_vocab"].get()
        dict = self.app_data["dict"]


        for widgets in self.frm_scroll.scrollable_frame.winfo_children():
            widgets.destroy()

        if var_vocab == 0:
            for key in dict:
                question_list.append(key)
        elif var_vocab == 1:
            for key in dict:
                question_list.append(dict[key])
        else:
            for key in dict:
                if random.randint(0, 1):
                    question_list.append(dict[key])
                else:
                    question_list.append(key)

        if self.app_data["word_shuffle"].get():
            random.shuffle(question_list)
        entries = self.app_data["entries"]
        entries.clear()
        for idx, text in enumerate(dict):
            if idx >= self.app_data["question_count"]:
                break
            label = ctk.CTkLabel(self.frm_scroll.scrollable_frame,width=150, text=question_list[idx])
            entry = ctk.CTkEntry(self.frm_scroll.scrollable_frame, width=200)
            label.grid(row=idx, column=0, sticky='we')
            entry.grid(row=idx, column=1, sticky='we')
            entries.append(entry)
            entry.bind("<Return>", self.nextWidget)
    def open_file(self):
         # 기존에 있던 값들 모두 초기화
        self.app_data["dict"].clear()
        self.app_data["word_count"] = 0
        self.textbox.configure(state='normal')
        self.textbox.delete(1.0, tk.END)

        fileName = filedialog.askopenfilenames(initialdir="/",\
            title = "파일을 선택 해 주세요", \
                filetypes = [("Excel Files", "*csv"), ("Text Files", "*txt"), ("All Files", "*.*")])
        if fileName == '' or fileName is None:
            return
        ext = os.path.splitext(fileName[0])[1]
        if ext == '.txt':
            with open(fileName[0], 'r', encoding='UTF-8') as file:
                lines = file.readlines()
                for line in lines:
                    pair = line.split(' ')
                    self.app_data["dict"][pair[0]] = pair[1].strip()
                    self.app_data["double_dict"][pair[0]] = pair[1].strip()
                    self.app_data["double_dict"][pair[1].strip()] = pair[0]
        elif ext == '.csv':
            with open(fileName[0], 'r') as file:
                self.app_data["dict"] = dict(csv.reader(file))
            for key,value in self.app_data["dict"].items():
                self.app_data["double_dict"][key] = value
                self.app_data["double_dict"][value] = key
        else:
            return
        self.textbox.insert('end', "단어 목록\n")
        for key,value in self.app_data["dict"].items():
            self.app_data["word_count"] += 1
            self.textbox.insert('end', str(self.app_data["word_count"]) + ". " + key + ": " + value + '\n')
        self.app_data["question_count"] = self.app_data["word_count"]

        self.textbox.configure(state='disabled')

if __name__ == "__main__":
    app = App()
    app.mainloop()

